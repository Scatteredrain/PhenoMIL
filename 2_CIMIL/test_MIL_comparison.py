import os
import torch

import sys
import cv2
import torch.nn as nn
from datetime import datetime
import json
import pandas as pd
import torch.nn.functional as F
# 要添加的目录路径
custom_dir = "/mnt/data/yizhenyu/data/HP识别/workspace/MIL-HP_trial"

# 转换为绝对路径并标准化
custom_dir = os.path.abspath(custom_dir)

# 检查路径是否存在
if not os.path.exists(custom_dir):
    raise FileNotFoundError(f"目录不存在: {custom_dir}")

# 添加到 Python 路径
if custom_dir not in sys.path:
    sys.path.insert(0, custom_dir)  # 优先搜索
    
# import torch.utils.data as data
from openpyxl import load_workbook
from lib.HP_encoder_aggregator import MILModel_combine
from utils.utils import *
from sklearn.metrics import accuracy_score, recall_score, roc_auc_score
# from utils.dataloader_multicenter_addImgLabel import PolypDataset
from tqdm.notebook import tqdm

import copy

def create_metrics(methods,centers):
  base_metrics = {"correct": 0, "total": 0,"TP": 0, "FP": 0,"TN": 0,"FN": 0,"all_labels": [],"all_probs": []}
  epoch_loss={"total_loss": 0.0, "img_loss": 0.0, "pat_loss": 0.0, "pat_center_loss": 0.0, "img_center_loss": 0.0}
  all_bag_feature = []
  multi_center_metrics = {
      center: {
          method: copy.deepcopy(base_metrics)
          for method in methods
      }
      for center in centers
  }
  return epoch_loss, all_bag_feature, multi_center_metrics

# 使用示例：添加数据到特定中心的方法
def update_metrics(multi_center_metrics, methods, centers, center_name, img_name, preds_patient,label_patient,probs_patient,labels_patient,split_shengzhongyi=False):
  for i, method_name in enumerate(methods):
    tp = ((preds_patient[i] == 1) & (label_patient == 1)).sum().item()
    fp = ((preds_patient[i] == 1) & (label_patient != 1)).sum().item()
    tn = ((preds_patient[i] != 1) & (label_patient != 1)).sum().item()
    fn = ((preds_patient[i] != 1) & (label_patient == 1)).sum().item()
    if center_name == 'HP_woNBI':
      center_name = '省中医'
    if center_name == '省中医' and split_shengzhongyi:
      # ['湖滨','城西'or'西溪','钱塘' or '下沙']
      if '湖滨' in img_name:
          center_yuanqu_name = '省中医-湖滨'
      elif '城西' in img_name or '西溪' in img_name:
          center_yuanqu_name = '省中医-城西'
      elif '钱塘' in img_name or '下沙' in img_name:
          center_yuanqu_name = '省中医-钱塘' 
      else:
          center_yuanqu_name = '省中医-湖滨'
          print('this center yuanqu belong to 省中医-湖滨', center_name, img_name)
       

      multi_center_metrics[center_yuanqu_name][method_name]['TP'] += tp
      multi_center_metrics[center_yuanqu_name][method_name]['FP'] += fp
      multi_center_metrics[center_yuanqu_name][method_name]['TN'] += tn
      multi_center_metrics[center_yuanqu_name][method_name]['FN'] += fn
      multi_center_metrics[center_yuanqu_name][method_name]['total'] += (tp + fp + tn + fn)
      multi_center_metrics[center_yuanqu_name][method_name]['correct'] += (tp + tn)
      multi_center_metrics[center_yuanqu_name][method_name]['all_probs'].extend(probs_patient[i])
      multi_center_metrics[center_yuanqu_name][method_name]['all_labels'].extend(labels_patient)
    else:
      multi_center_metrics[center_name][method_name]['TP'] += tp
      multi_center_metrics[center_name][method_name]['FP'] += fp
      multi_center_metrics[center_name][method_name]['TN'] += tn
      multi_center_metrics[center_name][method_name]['FN'] += fn
      multi_center_metrics[center_name][method_name]['total'] += (tp + fp + tn + fn)
      multi_center_metrics[center_name][method_name]['correct'] += (tp + tn)
      multi_center_metrics[center_name][method_name]['all_probs'].extend(probs_patient[i])
      multi_center_metrics[center_name][method_name]['all_labels'].extend(labels_patient)

def aggregate_all_centers(multi_center_metrics, methods, centers,):
    all_center = multi_center_metrics["所有中心"]
    for center in centers[1:]:  # 跳过"所有中心"
        for method in methods:
            for key in ['TP', 'FP', 'TN', 'FN', 'all_probs', 'all_labels']:
                all_center[method][key] += multi_center_metrics[center][method][key]

    # 更新衍生字段
    for method in methods:
        m = all_center[method]
        m['total'] = m['TP'] + m['FP'] + m['TN'] + m['FN']
        m['correct'] = m['TP'] + m['TN']

def print_result2(prefix, dict):
    # acc = dict["correct"] / (dict["total"]) * 100 if dict["total"] > 0 else 0
    # recall = dict["TP"] / (dict["TP"] + dict["FN"]) * 100 if (dict["TP"] + dict["FN"]) > 0 else 0
    # pre = dict["TP"] / (dict["TP"] + dict["FP"]) * 100 if (dict["TP"] + dict["FP"]) > 0 else 0
    # npv = dict["TN"] / (dict["TN"] + dict["FN"]) * 100 if (dict["TN"] + dict["FN"]) > 0 else 0
    # spe = dict["TN"] / (dict["TN"] + dict["FP"]) * 100 if (dict["TN"] + dict["FP"]) > 0 else 0
    # f1 = 2 * recall * pre / (recall + pre + 1e-6)
    # print(len)

        # dict["all_labels"] = np.array([x.squeeze()[0] for x in dict["all_labels"]])
    # all_labels = np.array([x.squeeze() for x in dict["all_labels"]])
    all_labels = np.array(dict["all_labels"])
    all_probs = np.array(dict["all_probs"])
    auc = roc_auc_score(all_labels, all_probs) 
    all_preds = np.array(all_probs) > 0.5
    all_preds = all_preds.astype(int)
    # print(len(all_labels), all_labels[0], len(all_probs), all_probs[0], len(all_preds), all_preds[0])
    metrics = calculate_metrics(all_labels,all_preds,all_probs)
    if len(prefix) >= 16:
        prefix += "\t"
    else:
        prefix += "\t\t"
    print(prefix, "Total:{}, Corr:{}, TP:{}, FP:{}, TN:{}, FN:{}\n\t\t\tAcc: {:.4f}, Rec: {:.4f}, Pre: {:.4f}, Spe: {:.4f}, Auc: {:.4f}, F1: {:.4f}"
          .format(dict["total"], dict["correct"], dict["TP"], dict["FP"], dict["TN"], dict["FN"],
                  metrics['Acc'],metrics['Sen'], metrics['Pre'], metrics['Spe'], auc, metrics['F1']))
    # res = [dict["total"], dict["correct"], dict["TP"], dict["FP"], dict["TN"], dict["FN"],
    #        metrics['Acc'],metrics['Sen'], metrics['Pre'], metrics['Spe'], auc, metrics['F1']]
    res =  "Total:{}, Corr:{}, TP:{}, FP:{}, TN:{}, FN:{}\n\t\t\tAcc: {:.4f}, Rec: {:.4f}, Pre: {:.4f}, Spe: {:.4f}, Auc: {:.4f}, F1: {:.4f}".format(
          dict["total"], dict["correct"], dict["TP"], dict["FP"], dict["TN"], dict["FN"],
                  metrics['Acc'],metrics['Sen'], metrics['Pre'], metrics['Spe'], auc, metrics['F1'])
    values = [dict["total"], dict["correct"], dict["TP"], dict["FP"], dict["TN"], dict["FN"],
                  metrics['Acc'],metrics['Sen'], metrics['Pre'], metrics['Spe'], auc, metrics['F1']]
    return res, values
    # return [dict["total"], dict["correct"], dict["TP"], dict["FP"], dict["TN"], dict["FN"],
    #         acc, recall, pre, spe, f1, auc]

def get_pooling(path):
  if '_Mean'in path:
    return 'Mean'
  elif '_Max'in path.lower():
    return 'Max'
  elif '_AB' in path:
    return 'AB'
  elif '_Trans' in path:
    return 'Trans'
  elif '_HPMIL' in path:
    return 'HPMIL'
  elif '_IAM' in path or '_HAGMIL' in path:
    return 'IAM'
  elif '_DSMIL' in path:
    return 'DSMIL'
  elif '_LSTM' in path:
    return 'LSTM'
  else:
    print('get_pooling error', path)
    return None

import argparse
from utils.dataloader_multicenter_addImgLabel import PolypDataset_test
if __name__ == '__main__':

  parser = argparse.ArgumentParser()
  parser.add_argument('--resume', type=str, default=None)
  opt = parser.parse_args()
  print('start loading')
  resume = opt.resume
  index_root = "/mnt/data/yizhenyu/data/HP识别/workspace/Hp-BTMIL/configs/multi_center_training4/test_data.txt"
  test_dataset = PolypDataset_test(
    index_root=index_root,
    transform_list=None,
    is_transform=False, is_train=False, testsize=352)

  test_loader = torch.utils.data.DataLoader(dataset=test_dataset,
                                      batch_size=1,
                                      shuffle=False,
                                      sampler=None,
                                      num_workers=8,
                                      pin_memory=True,
                                      drop_last=True)
  print('pos num:', len(test_dataset.names_p))
  print('neg num:', len(test_dataset.names_n))
  print('Total num:', len(test_dataset.names))


  methods = ['MIL']
  centers = ["所有中心", '余姚市人民医院','嘉一','宁波大学附属第一医院','温附一新','湖州新数据','省中医']
  topk=7
  instance_batch_size=20
  step_iter = enumerate(test_loader, start=1)

  pooling = get_pooling(resume)
  model = MILModel_combine(is_center_loss=False, is_img_center_loss=False, \
    is_img_loss=False, instance_batch_size=20, \
        aggregator=pooling, MaskDrop_threshold=0.7)
  weights = torch.load(resume)
  for k in list(weights.keys()):
      if k.startswith('module.'):
          weights[k[7:]] = weights.pop(k)
  model.load_state_dict(weights, strict=False)
  model.cuda()

  print('pooling:', pooling)
  print('load model from {}'.format(resume))

  epoch_loss, all_bag_feature, multi_center_metrics_test = create_metrics(methods,centers)
  model.eval()
  with torch.no_grad():
    for i, pack in tqdm(step_iter):

      (bags, bag_label, names, center_now) = pack
      center_now = center_now[0]

      if i % (len(test_dataset.names)//20) == 0:
        print(f'{i}/{len(test_dataset.names)}', end=" ")
      
      # -----------------
      _, max_instances, _, _, _ = bags.size() # bags: b=1, k, c, w, h
      if max_instances < topk:
        bags = torch.cat((bags, torch.zeros(bags.size(0), topk - max_instances, bags.size(2), bags.size(3), bags.size(4))),
                            dim=1)
        labels = torch.cat((labels, -torch.ones(labels.size(0), topk - max_instances)), dim=1)
        names.append(["zeros"]*(topk - max_instances))

      ## feature aggregation and classification
      out = model(bags)
      bag_out = out['bag_out']
      # {'bag_out': out_aggregation, 'importance_scores': importance_scores, 'aggregation_feature': aggregation_feature}

      bag_out = F.softmax(bag_out, dim=1)
      label_bag = bag_label.to(bag_out.device) # b
      # bag_loss = criterion(bag_out, label_bag)

      label_patient = label_bag.cpu().numpy()[:,0]
      labels_patient = label_patient #[label_patient]
      preds_patient = [bag_out.argmax(dim=1).detach().cpu().numpy()]
      probs_patient = [bag_out[:, 1].detach().cpu().numpy()]# 先收集到CPU
      #------------------

      # total_step+=1
      update_metrics(multi_center_metrics_test, methods, centers, center_now, names[0][0], preds_patient,label_patient,probs_patient,labels_patient)

  if len(centers) > 1:
    aggregate_all_centers(multi_center_metrics_test, methods, centers)

  print(f"Val Result of {resume}: ")
  reses = []
  data = []
  for center in centers:
    for method in methods:
      # method =  'pat_img_mean_dict'
      print('center: ',center, 'method: ',method)
      # print('center: ',centers[0],method)
      if multi_center_metrics_test[center][method]["total"] == 0:
        print('no counts')
        continue    
      res, values = print_result2('->:', multi_center_metrics_test[center][method])
      reses.append([center,res])
      data.append([center] + values)
      # f.write(f"{resume}\n{center}\t{method}\n{res}\n")
  with open('/mnt/data/yizhenyu/data/HP识别/workspace/MIL-HP_trial/results/Results_new_1021.txt', 'a') as f:
    f.write(f"\n\n{resume}\n")
    for center, res in reses:
      f.write(center + '\n' + res+'\n')
  # save to csv
  # df = pd.DataFrame(reses, columns=['center', 'method'])
  # df.to_csv('/mnt/data/yizhenyu/data/HP识别/workspace/MIL-HP_trial/results/Results.csv', index=False)
  
  # columns = ["total", "correct", "TP", "FP", "TN", "FN", "acc", "recall", "pre", "spe", "f1", "auc"]
  # df = pd.DataFrame(data, columns=["center"] + columns)

  # excel_file = "/mnt/data/yizhenyu/data/HP识别/workspace/MIL-HP_trial/results/results.xlsx"

  # # === 1. 准备要写入的数据块 ===
  # # 方法名行（单列）
  # method_df = pd.DataFrame([resume], columns=["Method"])

  # # 列名行（完整列名，含 "center"）
  # header_df = pd.DataFrame(columns=["center"] + columns)

  # # 数据行（df 已包含 "center"）

  # # === 2. 写入逻辑 ===
  # if os.path.exists(excel_file):
  #   # 追加模式
  #   book = load_workbook(excel_file)
  #   writer = pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay')
  #   writer._book = book
  #   writer.sheets = {ws.title: ws for ws in book.worksheets}
    
  #   # 获取当前 active sheet 的最后一行（1-indexed）
  #   sheet = book.active
  #   last_row = sheet.max_row  # 1-indexed
    
  #   # 下一个写入起始行（0-indexed） = last_row（已有行数）
  #   startrow = last_row  # pandas 的 startrow 是 0-indexed，last_row 是 1-indexed → 正好对应下一行
    
  #   # 写入：空行 + 方法名 + 列名 + 数据
  #   # 空行：通过 startrow 自动实现（因 last_row 是末尾，startrow = last_row 就是新行）
  #   method_df.to_excel(writer, index=False, header=False, startrow=startrow, startcol=0)
  #   header_df.to_excel(writer, index=False, startrow=startrow + 1, startcol=0)
  #   df.to_excel(writer, index=False, startrow=startrow + 2, startcol=0)
    
  # else:
  #   # 新建文件
  #   writer = pd.ExcelWriter(excel_file, engine='openpyxl')
  #   method_df.to_excel(writer, index=False, startrow=0, startcol=0)
  #   header_df.to_excel(writer, index=False, startrow=1, startcol=0)
  #   df.to_excel(writer, index=False, startrow=2, startcol=0)

  # writer.close()